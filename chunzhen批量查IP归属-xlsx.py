# -*- coding: utf-8 -*-
'''
@fucntion: 使用纯真ip数据库qqwry.dat批量查询txt内ip的归属地
https://soapffz.com/archives/245/
纯真IP数据库解析qqwry.dat库文件。 QQWry IP数据库 纯真版收集了包括中国电信、中国移动、中国联通、长城宽带、聚友宽带等 ISP 的最新准确 IP 地址数据。IP数据库每5天更新一次，需要定期更新最新的IP数据库。
下载qqwry.dat的地址：https://raw.githubusercontent.com/out0fmemory/qqwry.dat/master/qqwry_lastest.dat
pip install分别安装qqwry和IPy库
将要查询的ip放到ip.txt并和qqwry.dat、.py文件在同一个目录即可运行
'''

'''
Excel表格式：
    A：ip  B：数量  C：国家  D：地区
    A1     B1      C1      D1       
   →A2                     从这里开始读IP，开始查询，查询写入第三、四列

   Sheet1                  默认从这里加载数据
'''


from qqwry import QQwry
from IPy import IP
from openpyxl import load_workbook
import os

# print(os.path.split(os.path.realpath(__file__))[0])

china3 = ['局域网','重庆市','北京市','上海市','天津市','四川省','河北省','河南省','山东省','云南省','广东省','黑龙江','湖南省','山西省','安徽省','陕西省','湖北省','内蒙古','辽宁省','江西省','江苏省','甘肃省','浙江省','贵州省','西藏自','吉林省','青海省','海南省']
china2 = ['广西','新疆','西藏','宁夏']
#china_all = ['香港','澳门','台湾省','重庆市','北京市','上海市','天津市','四川省','河北省','河南省','山东省','云南省','广东省','黑龙江','湖南省','山西省','广西壮','安徽省','新疆维','陕西省','湖北省','内蒙古','辽宁省','江西省','江苏省','甘肃省','浙江省','贵州省','福建省','西藏自','吉林省','青海省','海南省','宁夏回']

q = QQwry()
q.load_file(os.path.split(os.path.realpath(__file__))[0] + '/qqwry_lastest.dat')

excel_file_path = os.path.split(os.path.realpath(__file__))[0] + '/ip.xlsx'
# 默认可读写，若有需要可以指定write_only和read_only为True
wb = load_workbook(excel_file_path)
# # 获得所有sheet的名称
# print(wb.get_sheet_names())
# # 获得当前正在显示的sheet, 也可以用wb.get_active_sheet()
# sheet = wb.active 
# print(sheet.title)
# 根据sheet名字获得sheet
######这里指定sheet的名称，可以修改#######
a_sheet = wb.get_sheet_by_name('Sheet1')
# # 获得sheet名
# print(a_sheet.title)
# # 获得最大列和最大行
# print(a_sheet.max_row)
# print(a_sheet.max_column)
# 读取指定sheet的指定列的所有内容
first_column = a_sheet['A']
for x in range(len(first_column)):
    if x > 0:
        ip_Address = first_column[x].value
        try:
            address = q.lookup(ip_Address)
            # print(x, ip_Address)
        except Exception:
            print('ip = struct.unpack(">I", socket.inet_aton(ip_str.strip()))[0]')
        try:
            IP(ip_Address)
        except Exception:
            print("有不符合规范的IP地址，请检查后重新运行")
            print(ip_Address)
            # exit(0)
        print(x, ip_Address,type(address),address)
        if address[0][:2] in china2:
            a_sheet.cell(x+1, 3, '中国')
            a_sheet.cell(x+1, 4, address[0])
        elif address[0][:3] in china3:
            # print(address,type(address),'中国',address[0])
            # 写入数据，三个参数分别是行，列，值
            a_sheet.cell(x+1, 3, '中国')
            a_sheet.cell(x+1, 4, address[0])
        else:
            # print(address, type(address), address[0][:3])
            a_sheet.cell(x+1, 3, address[0])
    # if x > 10:
    #     break

# 保存Excel表修改
wb.save(excel_file_path)

print('OK!')